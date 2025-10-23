import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

CRITERIA = [
    'Organization (10)',
    'Clarity & Communication (10)',
    'Q&A (10)',
    'Overall Participation (10)'
]

MONDAY_FILES = {
    'Kiran V K': 'monday_batch_evaluation_KVK.csv',
    'Shilpa': 'monday_batch_evaluation_Shilpa.csv',
    'Usha K': 'monday_batch_evaluation_UK.csv'
}

FRIDAY_FILES = {
    'Kiran V K': 'friday_batch_evaluation_KVK.csv',
    'Shankar': 'friday_batch_evaluation_shankar.csv',
    'Usha K': 'friday_batch_evaluation_UK.csv'
}

def normalize_roll_no(roll):
    if pd.isna(roll) or roll == '' or roll == '-':
        return None
    try:
        roll_str = str(roll).strip()
        if roll_str.upper() in ['NA', 'N/A', 'NAN']:
            return None
        roll_float = float(roll_str)
        return str(int(roll_float))
    except (ValueError, TypeError):
        return str(roll).strip() if str(roll).strip() else None

def normalize_name(name):
    if pd.isna(name) or name == '' or name == '-':
        return None
    name_str = str(name).strip()
    name_str = ' '.join(name_str.split())
    return name_str.title()

def is_blank(val):
    if pd.isna(val):
        return True
    if isinstance(val, str):
        val_str = val.strip().upper()
        return val_str in ['', '-', 'NA', 'N/A', 'NAN']
    return False

def coerce_to_numeric(val):
    if is_blank(val):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def read_and_normalize_csv(file_path, evaluator_name):
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return pd.DataFrame()
    
    df = df[df['Student Name'].notna() & (df['Student Name'] != '')]
    df = df[~df['Student Name'].str.strip().str.startswith('Group', na=False)]
    
    df['Roll No Normalized'] = df['Roll No'].apply(normalize_roll_no)
    df['Student Name Normalized'] = df['Student Name'].apply(normalize_name)
    
    df = df[df['Roll No Normalized'].notna() | df['Student Name Normalized'].notna()]
    
    for criterion in CRITERIA:
        if criterion in df.columns:
            df[f'{criterion}__{evaluator_name}'] = df[criterion].apply(coerce_to_numeric)
        else:
            df[f'{criterion}__{evaluator_name}'] = None
    
    if 'Total (40)' in df.columns:
        df[f'Total (40)__{evaluator_name}'] = df['Total (40)'].apply(coerce_to_numeric)
    else:
        df[f'Total (40)__{evaluator_name}'] = None
    
    keep_cols = ['Roll No Normalized', 'Student Name Normalized', 'Project Group', 
                 'Presentation Title', 'Guide Name']
    keep_cols += [f'{c}__{evaluator_name}' for c in CRITERIA]
    keep_cols += [f'Total (40)__{evaluator_name}']
    
    df = df[[c for c in keep_cols if c in df.columns or c.endswith(f'__{evaluator_name}')]]
    
    return df

def merge_evaluators(file_dict, batch_name):
    dfs = []
    evaluators = []
    data_issues = []
    
    for evaluator, filename in file_dict.items():
        df = read_and_normalize_csv(filename, evaluator)
        if not df.empty:
            dfs.append(df)
            evaluators.append(evaluator)
    
    if not dfs:
        return pd.DataFrame(), evaluators, data_issues
    
    merged = dfs[0].copy()
    
    for df in dfs[1:]:
        merged = pd.merge(
            merged, df,
            on=['Roll No Normalized', 'Student Name Normalized'],
            how='outer',
            suffixes=('', '_dup')
        )
    
    for col in ['Project Group', 'Presentation Title', 'Guide Name']:
        dup_cols = [c for c in merged.columns if c.startswith(col) and c != col]
        if dup_cols:
            merged[col] = merged[[col] + dup_cols].bfill(axis=1).iloc[:, 0]
            merged = merged.drop(columns=dup_cols)
    
    name_dup_cols = [c for c in merged.columns if c.startswith('Student Name Normalized_')]
    if name_dup_cols:
        merged = merged.drop(columns=name_dup_cols)
    
    merged = merged.sort_values('Roll No Normalized', na_position='last')
    
    return merged, evaluators, data_issues

def create_excel_with_formulas(merged_df, evaluators, batch_name, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{batch_name} Batch"
    
    # Build header row
    headers = ['Roll No', 'Student Name', 'Project Group', 'Presentation Title']
    
    # Column mapping for formula references
    col_map = {}
    current_col = 5  # Start after the first 4 columns
    
    for criterion in CRITERIA:
        # Add evaluator columns
        for ev in evaluators:
            headers.append(f'{criterion} ({ev})')
            col_map[f'{criterion}__{ev}'] = current_col
            current_col += 1
        
        # Add average column
        headers.append(f'Avg {criterion}')
        col_map[f'Avg_{criterion}'] = current_col
        current_col += 1
        
        # Add delta column
        headers.append(f'Δ {criterion}')
        col_map[f'Delta_{criterion}'] = current_col
        current_col += 1
    
    # Add total columns
    for ev in evaluators:
        headers.append(f'Total (40) ({ev})')
        col_map[f'Total__{ev}'] = current_col
        current_col += 1
    
    headers.append('Consolidated Total (40)')
    col_map['Consolidated_Total'] = current_col
    current_col += 1
    
    headers.extend(['Guide Name', 'Flag', 'Match Quality'])
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write data rows with formulas
    for row_idx, (_, row) in enumerate(merged_df.iterrows(), 2):
        # Basic info
        ws.cell(row=row_idx, column=1, value=row['Roll No Normalized'])
        ws.cell(row=row_idx, column=2, value=row['Student Name Normalized'])
        ws.cell(row=row_idx, column=3, value=row.get('Project Group', ''))
        ws.cell(row=row_idx, column=4, value=row.get('Presentation Title', ''))
        
        # Track average column letters for consolidated total
        avg_col_letters = []
        
        # For each criterion
        for criterion in CRITERIA:
            evaluator_cols = []
            
            # Write evaluator scores
            for ev in evaluators:
                col_name = f'{criterion}__{ev}'
                col_idx = col_map[col_name]
                value = row.get(col_name)
                if value is not None and not pd.isna(value):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                evaluator_cols.append(get_column_letter(col_idx))
            
            # Average formula (blank-aware)
            avg_col_idx = col_map[f'Avg_{criterion}']
            avg_col_letter = get_column_letter(avg_col_idx)
            avg_col_letters.append(f'{avg_col_letter}{row_idx}')
            
            eval_range = ','.join([f'{col}{row_idx}' for col in evaluator_cols])
            avg_formula = f'=IF(COUNTA({eval_range})=0,"",AVERAGE({eval_range}))'
            ws.cell(row=row_idx, column=avg_col_idx, value=avg_formula)
            
            # Delta formula (variance)
            delta_col_idx = col_map[f'Delta_{criterion}']
            delta_formula = f'=IF(COUNTA({eval_range})<2,"",MAX({eval_range})-MIN({eval_range}))'
            ws.cell(row=row_idx, column=delta_col_idx, value=delta_formula)
        
        # Write evaluator totals
        for ev in evaluators:
            col_name = f'Total (40)__{ev}'
            col_idx = col_map[f'Total__{ev}']
            value = row.get(col_name)
            if value is not None and not pd.isna(value):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Consolidated Total formula (sum of averages)
        cons_total_col_idx = col_map['Consolidated_Total']
        cons_total_formula = f'=SUM({",".join(avg_col_letters)})'
        ws.cell(row=row_idx, column=cons_total_col_idx, value=cons_total_formula)
        
        # Guide Name
        ws.cell(row=row_idx, column=current_col, value=row.get('Guide Name', ''))
        
        # Flag formula (check if any delta > 2)
        delta_cols = [get_column_letter(col_map[f'Delta_{c}']) + str(row_idx) for c in CRITERIA]
        flag_formula = f'=IF(OR({",".join([f"{col}>2" for col in delta_cols])}),"check variance","")'
        ws.cell(row=row_idx, column=current_col + 1, value=flag_formula)
        
        # Match Quality
        match_quality = 'roll' if row['Roll No Normalized'] else 'name_fallback'
        ws.cell(row=row_idx, column=current_col + 2, value=match_quality)
    
    # Apply formatting
    ws.freeze_panes = ws['A2']
    ws.auto_filter.ref = ws.dimensions
    
    # Conditional formatting for variance (Delta > 2)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for criterion in CRITERIA:
        delta_col_idx = col_map[f'Delta_{criterion}']
        delta_col_letter = get_column_letter(delta_col_idx)
        for row_idx in range(2, len(merged_df) + 2):
            cell = ws[f'{delta_col_letter}{row_idx}']
            # Add conditional formatting rule
            cell.number_format = '0.00'
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Set number format for score columns
    for criterion in CRITERIA:
        for ev in evaluators:
            col_idx = col_map[f'{criterion}__{ev}']
            col_letter = get_column_letter(col_idx)
            for row_idx in range(2, len(merged_df) + 2):
                ws[f'{col_letter}{row_idx}'].number_format = '0.0'
        
        avg_col_idx = col_map[f'Avg_{criterion}']
        avg_col_letter = get_column_letter(avg_col_idx)
        for row_idx in range(2, len(merged_df) + 2):
            ws[f'{avg_col_letter}{row_idx}'].number_format = '0.00'
        
        delta_col_idx = col_map[f'Delta_{criterion}']
        delta_col_letter = get_column_letter(delta_col_idx)
        for row_idx in range(2, len(merged_df) + 2):
            ws[f'{delta_col_letter}{row_idx}'].number_format = '0.00'
    
    # Set number format for totals
    cons_total_col_letter = get_column_letter(col_map['Consolidated_Total'])
    for row_idx in range(2, len(merged_df) + 2):
        ws[f'{cons_total_col_letter}{row_idx}'].number_format = '0.00'
    
    wb.save(output_file)
    print(f"[OK] {batch_name} Excel file saved: {output_file}")
    return len(merged_df)

def main():
    print('=' * 80)
    print('CONSOLIDATED EVALUATION EXCEL GENERATOR (WITH FORMULAS)')
    print('=' * 80)
    print()
    
    print('Processing Monday Batch...')
    monday_merged, evaluators_monday, issues_monday = merge_evaluators(MONDAY_FILES, 'Monday')
    
    if not monday_merged.empty:
        student_count = create_excel_with_formulas(
            monday_merged, 
            evaluators_monday, 
            'Monday',
            'Consolidated_Evaluation_Monday.xlsx'
        )
        print(f"  [OK] Monday batch: {student_count} students, {len(evaluators_monday)} evaluators")
    else:
        print('  [WARN] Monday batch: No data')
    
    print()
    print('Processing Friday Batch...')
    friday_merged, evaluators_friday, issues_friday = merge_evaluators(FRIDAY_FILES, 'Friday')
    
    if not friday_merged.empty:
        student_count = create_excel_with_formulas(
            friday_merged,
            evaluators_friday,
            'Friday',
            'Consolidated_Evaluation_Friday.xlsx'
        )
        print(f"  [OK] Friday batch: {student_count} students, {len(evaluators_friday)} evaluators")
    else:
        print('  [WARN] Friday batch: No data')
    
    print()
    print('=' * 80)
    print('GENERATION COMPLETE')
    print('=' * 80)
    print()
    print('Output files:')
    print('  - Consolidated_Evaluation_Monday.xlsx (with auto-calculating formulas)')
    print('  - Consolidated_Evaluation_Friday.xlsx (with auto-calculating formulas)')
    print()
    print('FEATURES:')
    print('  * Average columns use AVERAGE() formula - auto-updates when you add values')
    print('  * Delta columns use MAX()-MIN() formula - shows variance between evaluators')
    print('  * Consolidated Total uses SUM() formula - adds all average criteria')
    print('  * Flag column auto-detects variance > 2 and shows "check variance"')
    print('  * All formulas handle blank cells properly (blanks are NOT counted as zero)')
    print()

if __name__ == '__main__':
    main()
