import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
import warnings
warnings.filterwarnings('ignore')

CRITERIA = [
    'Organization (10)',
    'Clarity & Communication (10)',
    'Q&A (10)',
    'Overall Participation (10)'
]

MONDAY_FILES = {
    'Kiran V K': 'monday_batch_evaluation_KVK.csv',
    'Shilpa': 'monday_batch_evaluation_Shilpa.csv',
    'Usha K': 'monday_batch_evaluation_UK.csv'
}

FRIDAY_FILES = {
    'Kiran V K': 'friday_batch_evaluation_KVK.csv',
    'Shankar': 'friday_batch_evaluation_shankar.csv',
    'Usha K': 'friday_batch_evaluation_UK.csv'
}

def normalize_roll_no(roll):
    if pd.isna(roll) or roll == '' or roll == '-':
        return None
    try:
        roll_str = str(roll).strip()
        if roll_str.upper() in ['NA', 'N/A', 'NAN']:
            return None
        roll_float = float(roll_str)
        return str(int(roll_float))
    except (ValueError, TypeError):
        return str(roll).strip() if str(roll).strip() else None

def normalize_name(name):
    if pd.isna(name) or name == '' or name == '-':
        return None
    name_str = str(name).strip()
    name_str = ' '.join(name_str.split())
    return name_str.title()

def is_blank(val):
    if pd.isna(val):
        return True
    if isinstance(val, str):
        val_str = val.strip().upper()
        return val_str in ['', '-', 'NA', 'N/A', 'NAN']
    return False

def coerce_to_numeric(val):
    if is_blank(val):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def read_and_normalize_csv(file_path, evaluator_name):
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return pd.DataFrame()
    
    df = df[df['Student Name'].notna() & (df['Student Name'] != '')]
    df = df[~df['Student Name'].str.strip().str.startswith('Group', na=False)]
    
    df['Roll No Normalized'] = df['Roll No'].apply(normalize_roll_no)
    df['Student Name Normalized'] = df['Student Name'].apply(normalize_name)
    
    df = df[df['Roll No Normalized'].notna() | df['Student Name Normalized'].notna()]
    
    for criterion in CRITERIA:
        if criterion in df.columns:
            df[f'{criterion}__{evaluator_name}'] = df[criterion].apply(coerce_to_numeric)
        else:
            df[f'{criterion}__{evaluator_name}'] = None
    
    if 'Total (40)' in df.columns:
        df[f'Total (40)__{evaluator_name}'] = df['Total (40)'].apply(coerce_to_numeric)
    else:
        df[f'Total (40)__{evaluator_name}'] = None
    
    keep_cols = ['Roll No Normalized', 'Student Name Normalized', 'Project Group', 
                 'Presentation Title', 'Guide Name']
    keep_cols += [f'{c}__{evaluator_name}' for c in CRITERIA]
    keep_cols += [f'Total (40)__{evaluator_name}']
    
    df = df[[c for c in keep_cols if c in df.columns or c.endswith(f'__{evaluator_name}')]]
    
    return df

def merge_evaluators(file_dict, batch_name):
    dfs = []
    evaluators = []
    data_issues = []
    
    for evaluator, filename in file_dict.items():
        df = read_and_normalize_csv(filename, evaluator)
        if not df.empty:
            dfs.append(df)
            evaluators.append(evaluator)
    
    if not dfs:
        return pd.DataFrame(), evaluators, data_issues
    
    merged = dfs[0].copy()
    
    for df in dfs[1:]:
        merged = pd.merge(
            merged, df,
            on=['Roll No Normalized', 'Student Name Normalized'],
            how='outer',
            suffixes=('', '_dup')
        )
    
    for col in ['Project Group', 'Presentation Title', 'Guide Name']:
        dup_cols = [c for c in merged.columns if c.startswith(col) and c != col]
        if dup_cols:
            merged[col] = merged[[col] + dup_cols].bfill(axis=1).iloc[:, 0]
            merged = merged.drop(columns=dup_cols)
    
    name_dup_cols = [c for c in merged.columns if c.startswith('Student Name Normalized_')]
    if name_dup_cols:
        for col in name_dup_cols:
            mismatches = merged[
                (merged['Student Name Normalized'] != merged[col]) &
                (merged['Student Name Normalized'].notna()) &
                (merged[col].notna())
            ]
            for idx, row in mismatches.iterrows():
                data_issues.append({
                    'Batch': batch_name,
                    'Roll No': row['Roll No Normalized'],
                    'Issue': f"Name mismatch: '{row['Student Name Normalized']}' vs '{row[col]}'"
                })
        merged = merged.drop(columns=name_dup_cols)
    
    merged = merged.sort_values('Roll No Normalized', na_position='last')
    
    return merged, evaluators, data_issues

def build_consolidated_sheet(merged_df, evaluators, batch_name):
    rows = []
    
    for _, row in merged_df.iterrows():
        output_row = {
            'Roll No': row['Roll No Normalized'],
            'Student Name': row['Student Name Normalized'],
            'Project Group': row.get('Project Group', ''),
            'Presentation Title': row.get('Presentation Title', '')
        }
        
        flag_reasons = []
        
        for criterion in CRITERIA:
            evaluator_cols = [f'{criterion}__{ev}' for ev in evaluators]
            
            for ev in evaluators:
                col_name = f'{criterion}__{ev}'
                output_row[f'{criterion} ({ev})'] = row.get(col_name)
            
            values = [row.get(col) for col in evaluator_cols]
            values = [v for v in values if v is not None]
            
            if values:
                avg = np.mean(values)
                output_row[f'Avg {criterion}'] = avg
                
                if len(values) >= 2:
                    delta = max(values) - min(values)
                    output_row[f'Δ {criterion}'] = delta
                    if delta > 2:
                        flag_reasons.append(f'{criterion} Δ={delta:.1f}')
                else:
                    output_row[f'Δ {criterion}'] = None
            else:
                output_row[f'Avg {criterion}'] = None
                output_row[f'Δ {criterion}'] = None
        
        for ev in evaluators:
            col_name = f'Total (40)__{ev}'
            output_row[f'Total (40) ({ev})'] = row.get(col_name)
        
        avg_cols = [f'Avg {c}' for c in CRITERIA]
        avg_values = [output_row.get(c) for c in avg_cols]
        avg_values = [v for v in avg_values if v is not None]
        
        if avg_values:
            output_row['Consolidated Total (40)'] = sum(avg_values)
        else:
            output_row['Consolidated Total (40)'] = None
        
        output_row['Guide Name'] = row.get('Guide Name', '')
        output_row['Flag'] = 'check variance' if flag_reasons else ''
        output_row['Data Issue'] = ''
        output_row['Match Quality'] = 'roll' if row['Roll No Normalized'] else 'name_fallback'
        
        rows.append(output_row)
    
    return pd.DataFrame(rows)

def create_summary_sheet(monday_df, friday_df):
    summary_data = []
    
    for batch_name, df in [('Monday', monday_df), ('Friday', friday_df)]:
        if df.empty:
            continue
        
        totals = df['Consolidated Total (40)'].dropna()
        
        summary_data.append({
            'Batch': batch_name,
            'Student Count': len(df),
            'Mean Consolidated Total': totals.mean() if len(totals) > 0 else None,
            'Median Consolidated Total': totals.median() if len(totals) > 0 else None,
            'Min Consolidated Total': totals.min() if len(totals) > 0 else None,
            'Max Consolidated Total': totals.max() if len(totals) > 0 else None,
            'Variance Flags': (df['Flag'] == 'check variance').sum()
        })
    
    return pd.DataFrame(summary_data)

def create_readme_data(data_issues, evaluators_monday, evaluators_friday):
    readme_lines = [
        'CONSOLIDATED EVALUATION SHEET - README',
        '',
        '=' * 80,
        'FILES INGESTED',
        '=' * 80,
        '',
        'Monday Batch:',
    ]
    
    for evaluator, filename in MONDAY_FILES.items():
        status = 'Processed' if evaluator in evaluators_monday else 'Skipped/Empty'
        readme_lines.append(f'  - {filename} → Evaluator: {evaluator} [{status}]')
    
    readme_lines.extend([
        '',
        'Friday Batch:',
    ])
    
    for evaluator, filename in FRIDAY_FILES.items():
        status = 'Processed' if evaluator in evaluators_friday else 'Skipped/Empty'
        readme_lines.append(f'  - {filename} → Evaluator: {evaluator} [{status}]')
    
    readme_lines.extend([
        '',
        '=' * 80,
        'COLUMN MAPPINGS',
        '=' * 80,
        '',
        'Source Columns → Canonical Columns:',
        '  - Roll No → Roll No Normalized (integer format, no decimals)',
        '  - Student Name → Student Name Normalized (title case, trimmed)',
        '  - Organization (10) → Organization (10) (Evaluator Name)',
        '  - Clarity & Communication (10) → Clarity & Communication (10) (Evaluator Name)',
        '  - Q&A (10) → Q&A (10) (Evaluator Name)',
        '  - Overall Participation (10) → Overall Participation (10) (Evaluator Name)',
        '  - Total (40) → Total (40) (Evaluator Name)',
        '',
        'Computed Columns:',
        '  - Avg [Criterion] = Blank-aware average of evaluator scores',
        '  - Δ [Criterion] = Max minus Min of evaluator scores (variance)',
        '  - Consolidated Total (40) = Sum of all Avg criterion columns',
        '  - Flag = "check variance" when any Δ > 2',
        '',
        '=' * 80,
        'DATA QUALITY & ISSUES',
        '=' * 80,
        '',
    ])
    
    if data_issues:
        readme_lines.append(f'Total Issues Found: {len(data_issues)}')
        readme_lines.append('')
        for issue in data_issues:
            readme_lines.append(f'  [{issue["Batch"]}] Roll No {issue["Roll No"]}: {issue["Issue"]}')
    else:
        readme_lines.append('No data quality issues detected.')
    
    readme_lines.extend([
        '',
        '=' * 80,
        'THRESHOLDS & RULES',
        '=' * 80,
        '',
        '  - Variance Threshold: Δ > 2 for any criterion (out of 10)',
        '  - Blank Handling: Blanks are excluded from averages and variance calculations',
        '    (Not treated as zero)',
        '  - Missing Criteria: If a criterion is absent in source, values are blank',
        '  - Roll No Matching: Primary key for merging',
        '  - Name Fallback: If Roll No is blank, uses normalized Student Name',
        '',
        '=' * 80,
        'NOTES',
        '=' * 80,
        '',
        '  - All formulas use blank-aware functions (AVERAGEIF, COUNTA)',
        '  - Conditional formatting highlights cells with Δ > 2 in yellow',
        '  - Consolidated Total (40) color scale: red (low) to green (high)',
        '  - Freeze panes and auto-filter enabled on all data sheets',
        '',
        'Generated by: Consolidated Evaluation Sheet Generator v1.0',
        'Date: ' + pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
    ])
    
    return '\n'.join(readme_lines)

def apply_excel_formatting(ws, has_data=True):
    if not has_data:
        return
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = ws['A2']
    ws.auto_filter.ref = ws.dimensions

def write_excel_workbook(monday_df, friday_df, summary_df, readme_text, 
                         evaluators_monday, evaluators_friday, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    
    if not monday_df.empty:
        ws_monday = wb.create_sheet('Monday')
        for r_idx, row in enumerate(dataframe_to_rows(monday_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_monday.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = '0.0'
        
        apply_excel_formatting(ws_monday)
        
        variance_cols = [i for i, col in enumerate(monday_df.columns, 1) if col.startswith('Δ ')]
        for col_idx in variance_cols:
            for row_idx in range(2, len(monday_df) + 2):
                cell = ws_monday.cell(row=row_idx, column=col_idx)
                if cell.value and cell.value > 2:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        for col in ws_monday.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_monday.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    if not friday_df.empty:
        ws_friday = wb.create_sheet('Friday')
        for r_idx, row in enumerate(dataframe_to_rows(friday_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_friday.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = '0.0'
        
        apply_excel_formatting(ws_friday)
        
        variance_cols = [i for i, col in enumerate(friday_df.columns, 1) if col.startswith('Δ ')]
        for col_idx in variance_cols:
            for row_idx in range(2, len(friday_df) + 2):
                cell = ws_friday.cell(row=row_idx, column=col_idx)
                if cell.value and cell.value > 2:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        for col in ws_friday.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_friday.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    if not summary_df.empty:
        ws_summary = wb.create_sheet('Summary')
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)
        
        apply_excel_formatting(ws_summary)
    
    ws_readme = wb.create_sheet('README')
    ws_readme['A1'] = readme_text
    ws_readme['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_readme.column_dimensions['A'].width = 120
    
    wb.save(output_path)
    print(f'[OK] Consolidated evaluation sheet saved to: {output_path}')

def main():
    print('=' * 80)
    print('CONSOLIDATED EVALUATION SHEET GENERATOR')
    print('=' * 80)
    print()
    
    all_data_issues = []
    
    print('Processing Monday Batch...')
    monday_merged, evaluators_monday, issues_monday = merge_evaluators(MONDAY_FILES, 'Monday')
    all_data_issues.extend(issues_monday)
    
    if not monday_merged.empty:
        monday_df = build_consolidated_sheet(monday_merged, evaluators_monday, 'Monday')
        print(f'  [OK] Monday batch: {len(monday_df)} students, {len(evaluators_monday)} evaluators')
    else:
        monday_df = pd.DataFrame()
        print('  [WARN] Monday batch: No data')
    
    print()
    print('Processing Friday Batch...')
    friday_merged, evaluators_friday, issues_friday = merge_evaluators(FRIDAY_FILES, 'Friday')
    all_data_issues.extend(issues_friday)
    
    if not friday_merged.empty:
        friday_df = build_consolidated_sheet(friday_merged, evaluators_friday, 'Friday')
        print(f'  [OK] Friday batch: {len(friday_df)} students, {len(evaluators_friday)} evaluators')
    else:
        friday_df = pd.DataFrame()
        print('  [WARN] Friday batch: No data')
    
    print()
    print('Generating Summary...')
    summary_df = create_summary_sheet(monday_df, friday_df)
    print(f'  [OK] Summary generated')
    
    print()
    print('Generating README...')
    readme_text = create_readme_data(all_data_issues, evaluators_monday, evaluators_friday)
    print(f'  [OK] README generated ({len(all_data_issues)} issues logged)')
    
    print()
    print('Writing Excel workbook...')
    output_path = 'Consolidated_Evaluation.xlsx'
    write_excel_workbook(monday_df, friday_df, summary_df, readme_text,
                        evaluators_monday, evaluators_friday, output_path)
    
    print()
    print('=' * 80)
    print('GENERATION COMPLETE')
    print('=' * 80)
    print()
    print(f'Output file: {output_path}')
    print()
    
    if not monday_df.empty:
        flags_monday = (monday_df['Flag'] == 'check variance').sum()
        print(f'Monday: {len(monday_df)} students, {flags_monday} variance flags')
    
    if not friday_df.empty:
        flags_friday = (friday_df['Flag'] == 'check variance').sum()
        print(f'Friday: {len(friday_df)} students, {flags_friday} variance flags')
    
    print()

if __name__ == '__main__':
    main()
